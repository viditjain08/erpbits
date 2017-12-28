''' Use this script to populate django database with courses and slots it
it supports two types of excel sheets multipage and single page and by default assumes single page'''


import openpyxl
import sys
import django
import os

# set up django to access database populated
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'erp.settings')
django.setup()


from main.models import slot
from django.conf import settings


# starts creating and storing slots as per given arguments
def main(arguments, wb):

	# constants to keep track of course, type of course being registered in slot
	kwargs = {
	"course1" : '',
	"name1" : '',
	"teacher1" : '',
	"day1" : '',
	"hour1" : 0,
	"stype1" : '',
	"room1" : 0,
	"sec1" : 0,
	"flag" : 0,
	}

	if len(arguments)==0 or arguments[0]=="singlepage":
		singlepage(wb.worksheets[0], **kwargs) # pass first sheet to single page extractor
	elif  arguments[0]=="multipage":
		multipage(wb.worksheets, **kwargs) # pass list of sheets to extractor
	else:
		singlepage(wb.worksheets[0], **kwargs)


# extract data from a single sheet of the goven excel file
def singlepage(sheet, **kwargs):
	for i in range(2, sheet.max_row):
		kwargs = slotter(sheet, i, **kwargs)
		print i

# extract data from the lost of sheets in the given excel file
def multipage(sheetlist, **kwargs):
	# extracting data by iterating through sheets
	for sh in sheetlist:
		for l in range(2,sh.max_row):
			kwargs = slotter(sh, l, **kwargs)
			print l

# extract days and hours from column L and M in excel sheet
def dayhour(days, hours):
	daydict = {'M':'0','T':'1','W':'2','Th':'3','F':'4','S':'5'}
	day = ''.join(daydict[abc] for abc in days.split())
	hour = int(''.join(str(hours).split()))
	return day, hour

# creates or edits slots for each row in the given excel sheet
def slotter(sheet, l, course1, name1, teacher1, day1, hour1, stype1, room1, sec1, flag):
	if sheet['B'+str(l)].value: # if column B is not empty it is the indicative of a new course
			s = slot(totalseats=5, availableseats=5)
			s.course = course1 = sheet['B'+str(l)].value
			s.name = name1 = sheet['C'+str(l)].value
			s.teacher = teacher1 = sheet['H'+str(l)].value
			s.sec = sec1 = int(sheet['G'+str(l)].value)
			flag = 0
			if sheet['I'+str(l)].value == None:
				s.stype = stype1 = 'PROJECT' # if room number missing in course first row assign project
				s.room = room1 = 0
				s.day = day1 = ''
				s.hour = hour1 = 0
			else:
				s.stype = stype1 = ''
				s.room = room1 = int(sheet['I'+str(l)].value)
				s.day, s.hour = day1, hour1 = dayhour(sheet['J'+str(l)].value, sheet['K'+str(l)].value)
			s.save()
	elif sheet['C'+str(l)].value and not sheet['B'+str(l)].value: # continue with same course as previously discovered
		s = slot(totalseats=5, availableseats=5)			
		s.course = course1
		s.name = name1
		s.teacher = teacher1 = sheet['H'+str(l)].value
		s.stype = stype1 = sheet['C'+str(l)].value # change type to Practical, Tutorial etc as discovered
		s.day, s.hour = day1, hour1 = dayhour(sheet['J'+str(l)].value, sheet['K'+str(l)].value)
		s.room = room1 = int(sheet['I'+str(l)].value)
		if not sheet['G'+str(l)].value:
			s.sec = sec1 = 1
		else:
			s.sec = sec1 = sheet['G'+str(l)].value
		s.save()
	elif not sheet['C'+str(l)].value and not sheet['G'+str(l)].value: # assign multiple teachers for same slot/section
		s = slot.objects.get(course=course1,name=name1,sec=sec1,stype=stype1)
		s.teacher = s.teacher + '/ ' + sheet['H'+str(l)].value
		s.save()
	elif not sheet['C'+str(l)].value and sheet['G'+str(l)].value: # create a new section for course
		s = slot(totalseats=5, availableseats=5)			
		s.course = course1
		s.name = name1
		s.teacher = teacher1 = sheet['H'+str(l)].value
		s.stype = stype1
		s.day, s.hour = day1, hour1 = dayhour(sheet['J'+str(l)].value, sheet['K'+str(l)].value)
		s.room = room1 = int(sheet['I'+str(l)].value)
		s.sec = sec1 = sheet['G'+str(l)].value
		s.save()

	# return changed state variables back to kwargs to be passed a new to the function
	return {
	"course1" : course1,
	"name1" : name1,
	"teacher1" : teacher1,
	"day1" : day1,
	"hour1" : hour1,
	"stype1" : stype1,
	"room1" : room1,
	"sec1" : sec1,
	"flag" : flag,
	}

if __name__ == "__main__":
	arguments = sys.argv
	del arguments[0]
	wb = openpyxl.load_workbook('media/TIMETABLEII.xlsx')
	main(arguments, wb)